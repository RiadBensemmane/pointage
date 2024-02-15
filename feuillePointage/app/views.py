from django.shortcuts import render, redirect
from .models import Employe
from .forms import EmployeForm, EmployeFormUpdate
from datetime import datetime
from openpyxl import load_workbook
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
import os
# Create your views here.


def list(request):

    d, m, y = dmy(datetime.now())
    emp = Employe.objects.all()
    for e in emp:
        d2, m2, y2 = dmy(e.dateDernierPointage)
        if e.dateDernierPointage != None and (y2 != y or m2 != m or d2 != d):
            e.dateDernierPointage = None
            e.save(call=False)
            
        if e.dateDernierPointage == None:
            e.dernierPointage = None
            e.save(call=False)

    context = {'employes': emp, 'date': datetime.now()}
    
    return render(request, 'list.html', context)


def detail(request, pk):
    emp = Employe.objects.get(matricule=pk)
    context = {'employe': emp}
    return render(request, 'details.html', context)


def updateSheet(request, pk):

    pointage = request.POST.get('pointage')

    if pointage is not None:

        row_sheet_map, col_sheet_map = row_col_dic()    

        d, m, y = dmy(datetime.now())

        emp = Employe.objects.get(matricule=pk)

        wb = load_workbook('./PointageWorkbook/PointageAnnuel.xlsx')
        ws = wb[f"{emp.matricule} {emp.nom} {emp.prenom}"]

        ws[str(col_sheet_map[d]) + str(row_sheet_map[m])].value = pointage


        wb.save('./PointageWorkbook/PointageAnnuel.xlsx')

        emp.dernierPointage = pointage
        emp.dateDernierPointage = datetime.now()
        emp.save(call=False)
        
    return redirect('employe-list')


def ajouter_employe(request):
    if request.method == 'POST':
        form = EmployeForm(request.POST)
        if form.is_valid():
            form.save()
            msg = 'Nouvel employe ajouté avec succès'
            form = EmployeForm()
            return render(request, 'ajouter.html', {'form': form, 'message': msg})
        else:
            form = EmployeForm()
            msg = 'Veuillez au moins remplir les champs suivants: matricule, nom et prenom'
            return render(request, 'ajouter.html', {'form': form, 'message': msg})
    else:
        form = EmployeForm()
        msg = 'Veuillez au moins remplir les champs suivants: matricule, nom et prenom'
        return render(request, 'ajouter.html', {'form': form, 'message': msg})
    
def update_employe(request, pk):
    emp = Employe.objects.get(matricule=pk)

    matricule = emp.matricule
    nom = emp.nom
    prenom = emp.prenom
    if request.method == 'POST':
        form = EmployeFormUpdate(request.POST, instance=emp)
        if form.is_valid():
            change_sheet_values(matricule, nom, prenom, request.POST)
            form.save()
            return redirect('employe-detail', pk=pk)
        else:
            form = EmployeFormUpdate()
            msg = 'Veuillez remplir les champs'
            return render(request, 'update.html', {'form': form, 'message': msg, 'matricule': pk})
    
    else:
        form = EmployeFormUpdate(instance=emp)
        msg = 'Veuillez remplir les champs'
        return render(request, 'update.html', {'form': form, 'message': msg, 'matricule': pk})


def download(request):
    # Your path to the PointageWorkbook folder
    folder_path = './PointageWorkbook/'

    # Your Excel workbook name
    excel_file_name = 'PointageAnnuel.xlsx'

    # Full path to the Excel workbook
    excel_file_path = os.path.join(folder_path, excel_file_name)

    # Open the file for reading
    with open(excel_file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={excel_file_name}'
        return response



def row_col_dic():

    row_sheet_map = {i: i + 13 for i in range (1, 13)}

    col_sheet_map = {i: chr(i + ord('B') - 1 ) for i in range(1, 26)}
    dic2 = {i: 'A' + chr(i+39) for i in range(26, 32)}
    col_sheet_map.update(dic2)

    return row_sheet_map, col_sheet_map

def dmy(date):

    if date == None:
        return None, None, None
    
    y = date.year
    m = date.month
    d = date.day
    return d, m ,y

def change_sheet_values(matricule, nom, prenom, formAnswers):

    wb = load_workbook('./PointageWorkbook/PointageAnnuel.xlsx')
    ws = wb[f"{matricule} {nom} {prenom}"]

    ws.title = f"{matricule} {formAnswers['nom']} {formAnswers['prenom']}"


    ws['b6'].value = formAnswers['nom'] if ws['b6'].value != formAnswers['nom'] else ws['b6'].value
    ws['b7'].value = formAnswers['prenom'] if ws['b7'].value != formAnswers['prenom'] else ws['b7'].value
    ws['b8'].value = formAnswers['fonction'] if ws['b8'].value != formAnswers['fonction'] else ws['b8'].value
    ws['b10'].value = formAnswers['adresse'] if ws['b10'].value != formAnswers['adresse'] else ws['b10'].value
    ws['ag6'].value = formAnswers['dateRecrutement'] if ws['ag6'].value != formAnswers['dateRecrutement'] else ws['ag6'].value
    ws['ag8'].value = formAnswers['dateDetachement'] if ws['ag8'].value != formAnswers['dateDetachement'] else ws['ag8'].value
    ws['ag9'].value = formAnswers['affectOrigine'] if ws['ag9'].value != formAnswers['affectOrigine'] else ws['ag9'].value
    ws['ag10'].value = formAnswers['situationFamiliale'] if ws['ag10'].value != formAnswers['situationFamiliale'] else ws['ag10'].value
    ws['ag11'].value = formAnswers['nbrEnfants'] if ws['ag11'].value != formAnswers['nbrEnfants'] else ws['ag11'].value

    wb.save('./PointageWorkbook/PointageAnnuel.xlsx')   
