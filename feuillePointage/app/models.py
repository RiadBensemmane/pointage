from django.db import models
from django.core.validators import FileExtensionValidator
from openpyxl import load_workbook
# Create your models here.


class Employe(models.Model):
    CHOICES = {
        ('marié', 'marié(e)'),
        ('divorcé', 'divorcé(e)'),
        ('célibataire', 'célibataire'),
        ('veuf', 'veuf(ve)')
    }
    CODES = {
        ('T', 'T'), ('R', 'R'), ('1', '1'), ('I', 'I'), ('2', '2'), ('M', 'M'), ('A', 'A'), ('6', '6'), ('7', '7'),
        ('8', '8'), ('9', '9'), ('C', 'C'), ('Cr', 'Cr'),  
    }
    matricule = models.PositiveIntegerField(primary_key = True)
    nom = models.CharField(max_length = 100)
    prenom = models.CharField(max_length = 100)
    adresse = models.CharField(max_length = 200, null=True)
    fonction = models.CharField(max_length = 100, null=True)
    dateRecrutement = models.DateField(null=True)
    dateDetachement = models.DateField(null=True)
    affectOrigine = models.CharField(max_length = 100, null=True)
    situationFamiliale = models.CharField(max_length = 100, choices= CHOICES, null=True)
    nbrEnfants = models.PositiveIntegerField(null=True)
    dernierPointage = models.CharField(max_length = 2, choices=CODES, default=None, null=True)
    dateDernierPointage = models.DateField(null=True, default = None)

    def __str__(self):
        return self.nom + ' ' + self.prenom
    
    def save(self, *args, **kwargs):
        
        call_parameter = kwargs.pop('call', True)  # Default value is True if not provided

        # Your custom save logic here, if needed

        super(Employe, self).save(*args, **kwargs)

        if call_parameter:
            self.create_employe_sheet()



    def create_employe_sheet(self): 

        wb = load_workbook('./PointageWorkbook/PointageAnnuel.xlsx')

        sourceSheet = wb['template']

        title = f"{self.matricule} {self.nom} {self.prenom}"

        if title not in wb.sheetnames:
            newsheet = wb.copy_worksheet(sourceSheet)
            newsheet.title = title
            newsheet['b6'].value = self.nom
            newsheet['b7'].value = self.prenom
            newsheet['b8'].value = self.fonction
            newsheet['b10'].value = self.adresse
            newsheet['u6'].value = self.matricule
            newsheet['ag6'].value = self.dateRecrutement
            newsheet['ag9'].value = self.affectOrigine
            newsheet['ag10'].value = self.situationFamiliale
            newsheet['ag11'].value = self.nbrEnfants

        wb.save('./PointageWorkbook/PointageAnnuel.xlsx')
            

 